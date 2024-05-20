from django.shortcuts import render
from openpyxl.reader.excel import load_workbook
from django.contrib.auth.decorators import login_required
from .models import Application, ClientInfo, CreditPay, CreditPayment, AddClients
from rest_framework.views import APIView
from rest_framework.response import Response
from .serializers import AppSerializer, ClientInfoSerializer, CreditPaySerializer, CreditPaymentSerializer
from django.http import JsonResponse
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.shortcuts import redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from .forms import ApplicationSearchForm
from django.contrib import messages
# import datetime
from datetime import datetime

# Create your views here.

now = datetime.now()
dt_string = now.strftime("%d-%m-%Y %H:%M:%S")
date = str(dt_string)


def base(request):
    return render(request, 'base.html')

@login_required()
def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file.name.endswith('.xlsx'):
            wb = load_workbook(excel_file)
            ws = wb.active
            instances = []
            Application.objects.all().delete()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row:
                    try:
                        change_date = datetime.strptime(row[4], '%d.%m.%Y').date()
                        credit_date = datetime.strptime(row[12], '%d.%m.%Y').date()
                        credit_term_date = datetime.strptime(row[13], '%d.%m.%Y').date()
                        first_pay_date = datetime.strptime(row[18], '%d.%m.%Y').date()
                        last_pay_date = datetime.strptime(row[19], '%d.%m.%Y').date()

                        a = Application.objects.create(
                            mfo=row[0],
                            branch_id=row[1],
                            application_id=row[2],
                            state=row[3],
                            change_date=change_date,
                            loan_id=row[5],
                            region_code=row[6],
                            district_code=row[7],
                            client_type=row[8],
                            client_name=row[9],
                            client_id=row[10],
                            credit_num=row[11],
                            credit_date=credit_date,
                            credit_term_date=credit_term_date,
                            credit_sum=row[14],
                            credit_percent=row[15],
                            credit_account=row[16],
                            total_pay_sum=row[17],
                            first_pay_date=first_pay_date,
                            last_pay_date=last_pay_date,
                            debt_sum=row[20],
                            overdue_debt_sum=row[21],
                            credit_purpose_code=row[22],
                            credit_purpose_name=row[23],
                            account_16309_sum=row[24],
                            account_16323_sum=row[25],
                            account_16377_sum=row[26],
                            account_16379_sum=row[27],
                        )
                        instances.append(a)
                    except Exception as e:
                        print(f"Error processing row {row}: {e}")
                else:
                    messages.info(request, 'Your rows are not correspond to save it')
            if instances:
                messages.success(request, 'Your file successfully loaded')
            else:
                messages.info(request, 'Your rows are not correspond to save it')
        else:
            messages.warning(request, 'You must only excel file upload')
    return render(request, 'upload_file.html')


@login_required()
def credit_info_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file.name.endswith('.xlsx'):
            wb = load_workbook(excel_file)
            ws = wb.active
            instances = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if CreditPay.objects.filter(loan_id=row[0]).delete():
                    try:
                        a = CreditPay.objects.create(
                            loan_id=1,
                            mfo=row[0],
                            account=row[1],
                            nibbd_code=row[2],
                            branch_id=row[3],
                            account_status=row[4],
                            turnover_db_20208=row[5],
                            turnover_cr_20208=row[6],
                            turnover_db_20212=row[7],
                            turnover_cr_20212=row[8],
                            turnover_cr_20218=row[9],
                            turnover_db_20218=row[10],
                            saldo_90963=row[11]
                        )
                        instances.append(a)
                    except Exception as e:
                        print(f"Error processing row {row}: {e}")
                else:
                    messages.info(request, 'Your rows are not correspond to save it')
            if instances:
                messages.success(request, 'Your file successfully loaded')
            else:
                messages.info(request, 'Your rows are not correspond to save it')
        else:
            messages.warning(request, 'You must only excel file upload')
    return render(request, 'upload_file.html')


import requests
from django.http import HttpResponse


def checkclients(request):
    url = "http://10.9.49.250:80/api/v1/bank/applications?last_id=0"

    headers = {
        'Authorization': 'Basic dHVyb25iYW5rOkVpZTZHWWVROUtoUGU2NDQ='
    }

    response = requests.request("GET", url, headers=headers)

    data = response.json()

    a = Application.objects.values_list('client_id', flat=True)
    data1 = list(a)

    for key in data['data']:
        if key['doc_status'] == 'ACCEPTED':
            Application.objects.filter(client_id=key['client_id']).update(loan_id=key['doc_num'])

    return HttpResponse('success')


def addclients(request):
    url = "http://10.9.49.250:80/api/v1/bank/applications?last_id=0"

    headers = {
        'Authorization': 'Basic dHVyb25iYW5rOkVpZTZHWWVROUtoUGU2NDQ='
    }

    response = requests.request("GET", url, headers=headers)

    data = response.json()

    instances = []

    AddClients.objects.all().delete()
    for key in data['data']:
        if key['doc_status'] == 'ACCEPTED':
            a = AddClients.objects.create(
                client_type=key['client_type'],
                client_id=key['client_id'],
                client_name=key['client_name'],
                nibbd_code=key['nibbd_code'],
                doc_date=datetime.strptime(key['doc_date'], '%d.%m.%Y'),
                doc_status=key['doc_status'],
                doc_guid=key['doc_guid'],
                doc_pin=key['doc_pin'],
                doc_num=key['doc_num'],
                credit_sum=key['credit_sum'],
                credit_term=key['credit_term'],
                credit_privilege_month=key['credit_privilege_month'],
                credit_percent=key['credit_percent'],
                workers=key['workers'],
                credit_type=key['credit_type'],
                region_code=key['region_code'],
                district_code=key['district_code'],
                branch_id=key['branch_id'],
                update_date=datetime.strptime(key['update_date'], '%d.%m.%Y'),
                update_id=key['update_id']
            )
            instances.append(a)
            # if instances:
    messages.success(request, "Mijozlar ro'yxati yangilandi")
            # else:
            #     messages.info(request, "Mijoz ro'yxatini yangilashda xatolik mavjud")

    return render(request, 'home.html')


def checkpayment(request):
    url = "http://10.9.49.250:80/api/v1/bank/applications?last_id=0"

    headers = {
        'Authorization': 'Basic dHVyb25iYW5rOkVpZTZHWWVROUtoUGU2NDQ='
    }

    response = requests.request("GET", url, headers=headers)

    data = response.json()

    a = CreditPayment.objects.values_list('ca_id', flat=True)
    data1 = list(a)

    for key in data['data']:
        if key['doc_status'] == 'ACCEPTED':
            CreditPayment.objects.filter(ca_id=key['client_id']).update(loan_id=key['doc_num'])

    return HttpResponse('success')


@login_required()
def credit_payment_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file.name.endswith('.xlsx'):
            wb = load_workbook(excel_file)
            ws = wb.active
            instances = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row:
                    try:
                        doc_date = datetime.strptime(row[10], '%d.%m.%Y')
                        credit_date = datetime.strptime(row[16], '%d.%m.%Y')

                        a = CreditPayment.objects.create(
                            branch_id=row[0],
                            cl_mfo=row[1],
                            cl_account=row[2],
                            cl_name=row[3],
                            cl_id=row[4],
                            ca_mfo=row[5],
                            ca_account=row[6],
                            ca_name=row[7],
                            ca_id=row[8],
                            doc_id=row[9],
                            doc_date=doc_date,
                            doc_num=row[11],
                            pay_sum=row[12],
                            pay_code=row[13],
                            pay_note=row[14],
                            pay_state=row[15],
                            pay_date=credit_date,
                            loan_id=row[17]
                        )
                        instances.append(a)
                    except Exception as e:
                        print(f"Error processing row {row}: {e}")
                else:
                    messages.info(request, 'Your rows are not correspond to save it')
            if instances:
                messages.success(request, 'Your file successfully loaded')
            else:
                messages.info(request, 'Your rows are not correspond to save it')
        else:
            messages.warning(request, 'You must only excel file upload')
    return render(request, 'upload_file.html')


class PaymentView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        request_id = request.query_params.get('request_id')
        loan_id = request.query_params.get('loan_id')
        brb = request.query_params.get('date')
        changed_date = datetime.strptime(brb, '%d.%m.%Y').strftime("%Y-%m-%d")
        if CreditPayment.objects.filter(loan_id=loan_id).exists():
            try:
                # ClientInfo.objects.create(request_id=request_id, application_id=application_id)
                application_obj = CreditPayment.objects.filter(loan_id=loan_id, pay_date=changed_date)
                serializer = CreditPaymentSerializer(application_obj, many=True)
                return Response(serializer.data)
            except CreditPayment.DoesNotExist:
                return Response({'error': 'Object not found'}, status=status.HTTP_404_NOT_FOUND)
        elif not CreditPayment.objects.filter(loan_id=loan_id).exists():
            return Response({'error': '00001',
                             'message': "loan_id or date not found",
                             'timestamp': date,
                             'status': 404})


class ApplicationView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        request_id = request.query_params.get('request_id')
        application_id = request.query_params.get('application_id')
        if Application.objects.filter(loan_id=application_id).exists():
            try:
                ClientInfo.objects.create(request_id=request_id, application_id=application_id)
                application_obj = Application.objects.filter(loan_id=application_id)
                serializer = AppSerializer(application_obj, many=True)
                return Response(serializer.data)
            except Application.DoesNotExist:
                return Response({'error': 'Object not found'}, status=status.HTTP_404_NOT_FOUND)
        elif not Application.objects.filter(loan_id=application_id).exists():
            return Response({'error': '00001',
                             'message': "application_id not found",
                             'timestamp': date,
                             'status': 404})


class CreditView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        request_id = request.query_params.get('request_id')
        nibbd_code = request.query_params.get('nibbd_code')
        if CreditPay.objects.filter(nibbd_code=nibbd_code).exists():
            try:
                # CreditPay.objects.create(request_id=request_id, application_id=application_id)
                application_obj = CreditPay.objects.filter(nibbd_code=nibbd_code)
                serializer = CreditPaySerializer(application_obj, many=True)
                return Response(serializer.data)
            except CreditPay.DoesNotExist:
                return Response({'error': 'Object not found'}, status=status.HTTP_404_NOT_FOUND)
        elif not CreditPay.objects.filter(nibbd_code=nibbd_code).exists():
            return Response({'error': '00001',
                             'message': "nibbd_code not found",
                             'timestamp': date,
                             'status': 404})


def logout_view(request):
    if request.method == 'POST':
        logout(request)
        return redirect('login')


def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')
        else:
            return redirect('login')

    return render(request, 'login.html')


@login_required()
def home(request):
    items_per_page = 20

    page = request.GET.get('page')

    queryset = AddClients.objects.all()
    search_form = ApplicationSearchForm(request.GET)

    if search_form.is_valid():
        loan_id = search_form.cleaned_data.get('loan_id')

        if loan_id:
            queryset = queryset.filter(loan_id__icontains=loan_id)

    paginator = Paginator(queryset, items_per_page)

    try:
        current_page = paginator.page(page)
    except PageNotAnInteger:
        current_page = paginator.page(1)
    except EmptyPage:
        current_page = paginator.page(paginator.num_pages)
    num_pages = paginator.num_pages
    current_page_number = current_page.number
    page_range = list(range(max(current_page_number - 2, 1), min(current_page_number + 3, num_pages + 1)))

    if page_range[0] > 2:
        page_range.insert(0, 1)
        page_range.insert(1, None)

    if page_range[-1] < num_pages - 1:
        page_range.append(None)
        page_range.append(num_pages)

    return render(request, 'home.html',
                  {'current_page': current_page, 'search_form': search_form, 'page_range': page_range})


import oracledb


def do(self):
    sql_request = f"""
    --call c_pkgconnect.popen()
select 
--|| ' ' || j.dep_id as doc_id,
j.code_bcl as branch_id,
j.code_bcl as cl_mfo,
a.code as cl_account,
j.TXT_PAY as cl_name,
decode(j.INCOMFL, '1', j.RNN_CR, j.RNN_CL) as cl_id, 
       j.code_bcr as ca_mfo,
              decode(j.INCOMFL, '1', j.CODE_ACL, j.CODE_ACR) as ca_account,
             j.txt_ben as ca_name,
             decode(j.INCOMFL, '1', j.RNN_CL, j.RNN_CR) as ca_id,
              j.id  as doc_id,
             t.doper as doc_date,
             ord.CODE as doc_num,
             t.sdok as pay_sum,
             j.knp as pay_code,
       --case t.incomfl
       --  when '1' then
       --    'dt'
       --   else
       --   'kt'
       --end,
       j.txt_dscr as pay_note,
       (case t.postfl
         when '0' then '00'
            when '1' then '41'
             else '13'
             end) as pay_state,
       t.doper as pay_date
  from g_accbln   a,
       ledacc_std l,
       t_trndtl   t,
       t_trndtl   t2,
       g_accbln   a2, 
       t_Operjrn jrn,
       S_ORDPAY   pay,
       p_ord j,
       t_ord ord
 where a.cha_id = l.id
   and t.dep_id = a.dep_id
   and t.acc_id = a.id
   and t2.dep_id = a2.dep_id
   and t2.acc_id = a2.id
   and t.ID = t2.ID(+)
   and t.NORD = t2.NORD(+)
   and t.INCOMFL != t2.INCOMFL(+)
   --and a.code ='15521000300283864981'
   
   and t.ID = jrn.TRA_ID(+)
   and jrn.DEP_ID = ord.DEP_ID(+)
   and jrn.ORD_ID = ord.ID(+)
      
   and ord.DEP_ID = pay.dep_id(+)
   and ord.ID = pay.ORD_ID(+)
   and ord.dep_id = j.dep_id(+)
   and ord.id = j.id(+)
   
   and l.code in ('15521', '15021')
   and j.code_acr like '202%'
   and t.incomfl = 1
   and t.doper between to_date('28.02.2024', 'dd.mm.yyyy') and to_date('15.05.2024', 'dd.mm.yyyy')
       """

    r = []
    if sql_request:
        p_username = "Colvir"
        p_password = "ColvirTuron"
        p_dns = "192.168.7.152/CBSPROD"
        p_port = "1521"

        with oracledb.connect(
                user=p_username, password=p_password, dsn=p_dns, port=p_port
        ) as connection:
            with connection.cursor() as cursor:
                cursor.execute('begin c_pkgconnect.popen(iNotRegist=>1); end;')
                cursor.execute(sql_request, {})
                column_names = [desc[0] for desc in cursor.description]
                rows = cursor.fetchall()
                result = []
                for row in rows:
                    result.append(dict(zip(column_names, row)))
                l = []
                CreditPayment.objects.all().delete()
                for i in result:
                    CreditPayment.objects.create(
                        branch_id=i['branch_id'],
                        loan_id=1,
                        cl_mfo=i['cl_mfo'],
                        cl_account=i['cl_account'],
                        cl_name=i['cl_name'],
                        cl_id=i['cl_id'],
                        ca_mfo=i['ca_mfo'],
                        ca_account=i['ca_account'],
                        ca_name=i['ca_name'],
                        ca_id=i['ca_id'],
                        doc_id=i['doc_id'],
                        doc_date=i['doc_date'],
                        doc_num=i['doc_num'],
                        pay_sum=i['pay_sum'],
                        pay_code=i['pay_code'],
                        pay_note=i['pay_note'],
                        pay_state=i['pay_state'],
                        pay_date=i['pay_date'],
                    )

        return HttpResponse('success')

# SERVER
# @login_required()
# def import_from_excel(request):
#     if request.method == 'POST':
#         excel_file = request.FILES.get('excel_file')
#         if excel_file.name.endswith('.xlsx'):
#             wb = load_workbook(excel_file)
#             ws = wb.active
#             instances = []
#             for row in ws.iter_rows(min_row=2, values_only=True):
#                 if not Application.objects.filter(loan_id=row[5]).exists():
#                     try:
#                         change_date = datetime.strptime(row[4], '%d.%m.%Y').date()
#                         credit_date = datetime.strptime(row[12], '%d.%m.%Y').date()
#                         credit_term_date = datetime.strptime(row[13], '%d.%m.%Y').date()
#                         first_pay_date = datetime.strptime(row[18], '%d.%m.%Y').date()
#                         last_pay_date = datetime.strptime(row[19], '%d.%m.%Y').date()
#
#                         a = Application.objects.create(
#                             mfo=row[0],
#                             branch_id=row[1],
#                             application_id=row[2],
#                             state=row[3],
#                             change_date=change_date,
#                             loan_id=row[5],
#                             # claim_id=row[5],
#                             region_code=row[6],
#                             district_code=row[7],
#                             client_type=row[8],
#                             client_name=row[9],
#                             client_id=row[10],
#                             credit_num=row[11],
#                             credit_date=credit_date,
#                             credit_term_date=credit_term_date,
#                             credit_sum=row[14],
#                             credit_percent=row[15],
#                             credit_account=row[16],
#                             total_pay_sum=row[17],
#                             first_pay_date=first_pay_date,
#                             last_pay_date=last_pay_date,
#                             debt_sum=row[20],
#                             overdue_debt_sum=row[21],
#                             credit_purpose_code=row[22],
#                             credit_purpose_name=row[23],
#                             account_16309_sum=row[24],
#                             account_16323_sum=row[25],
#                             account_16377_sum=row[26],
#                             account_16379_sum=row[27],
#                         )
#                         instances.append(a)
#                     except Exception as e:
#                         print(f"Error processing row {row}: {e}")
#                 else:
#                     messages.info(request, 'Your rows are not correspond to save it')
#             if instances:
#                 messages.success(request, 'Your file successfully loaded')
#             else:
#                 messages.info(request, 'Your rows are not correspond to save it')
#         else:
#             messages.warning(request, 'You must only excel file upload')
#     return render(request, 'upload_file.html')
